import openpyxl


workbook=openpyxl.load_workbook('./../data/Book2.xlsx')

sheet=workbook['Sheet1']
rc=sheet.max_row #3
print('Row Count',rc)
cc=sheet.max_column #4
print('Col Count',cc)

for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=sheet.cell(i,j).value
        print(v)


workbook.close()
