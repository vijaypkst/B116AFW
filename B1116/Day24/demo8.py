import time

import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By

driver=Chrome()
driver.implicitly_wait(10)
driver.get("https://pos.aksharatraining.in/pos/public")
driver.maximize_window()

workbook=openpyxl.load_workbook("./../data/pos.xlsx")
un=workbook['login'].cell(2,1).value
pw=workbook['login'].cell(2,2).value
expected_url=workbook['login'].cell(2,3).value

driver.find_element(By.ID,"input-username").send_keys(un)
driver.find_element(By.ID,"input-password").send_keys(pw)
driver.find_element(By.NAME,"login-button").click()

time.sleep(3)
actual_url=driver.current_url
workbook['login'].cell(2,4).value=actual_url

if actual_url==expected_url:
    workbook['login'].cell(2, 5).value ="PASS"
else:
    workbook['login'].cell(2, 5).value = "FAIL"

workbook.save("./../data/pos_result.xlsx")
workbook.close()
driver.close()